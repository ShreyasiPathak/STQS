# -*- coding: utf-8 -*-
"""
Created on Fri Mar  5 19:03:33 2021

@author: PathakS
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 18:28:49 2019

@author: PathakS
"""

import os
import h5py
import torch
import pickle
import random
import argparse
import numpy as np
import torch.nn as nn
import openpyxl as op
from random import shuffle
import torch.optim as optim
from openpyxl import Workbook
import matplotlib.pyplot as plt
from collections import OrderedDict
from pytorchtools import EarlyStopping
from torch.utils.data import DataLoader, SequentialSampler
from sklearn.metrics import confusion_matrix

import utils
import models

def args_parameters():
    parser = argparse.ArgumentParser()
    parser.add_argument('--batch_size', type=int, default=192)
    parser.add_argument('--n_workers', type=int, default=16)
    parser.add_argument('--learning_rate', type=float, default=0.0001)
    parser.add_argument('--max_epochs', type=int, default=200)
    parser.add_argument('--time_steps', type=int, default=3750)
    parser.add_argument('--n_channels', type=int, default=5)
    parser.add_argument('--modality_pipelines', type=int, default=3)
    parser.add_argument('--seq_length', type=int, default=8)
    parser.add_argument('--class_imbalance', type=str, default='oversampling')
    parser.add_argument('--lstm_option', type=bool, default=False)
    parser.add_argument('--rc_option', type=bool, default=False)
    parser.add_argument('--patience_epoch', type=int, default=7)
    args = parser.parse_args()
    return args

def load_trained_model_for_testing(model, path):
    model_parameters=torch.load(path)
    model_weights=model_parameters['state_dict']
    state_dict_remove_module = OrderedDict()
    for k, v in model_weights.items():
        #print(k, v.shape) #k coming from pretrained model
        name = k[7:] # remove `module.`
        state_dict_remove_module[name] = v
    #state_dict_new_model=model.state_dict()
    #print(state_dict_new_model.keys())
    #model.load_state_dict(model_weights) #new model getting updated with the trained weights
    model.load_state_dict(state_dict_remove_module)
    #print(model.channelBlocks['eeg'].layer1[0].weight)
    return model.to(device)

def load_pretrained_model_for_LSTM(model,path):
    state_dict_new_model=model.state_dict()
    checkpoint = torch.load(path)#, map_location=device)
    state_dict_pretrained=checkpoint['state_dict']
    state_dict_remove_module = OrderedDict()
    for k, v in state_dict_pretrained.items():
        if k[:7]=='module.':
            if k!='module.linear.weight' and k!='module.linear.bias':
                if device=="cuda:0":
                    state_dict_remove_module[k] = v
                else:
                    name = k[7:] # remove `module.`
                    state_dict_remove_module[name] = v
        else:
            if k!='linear.weight' and k!='linear.bias':
                if device=="cuda:0":
                    name = 'module.' + k
                    state_dict_remove_module[name] = v
                else:
                    state_dict_remove_module[k] = v
    state_dict_new_model.update(state_dict_remove_module)
    model.load_state_dict(state_dict_new_model)
    return model

def save_model(model,optimizer,epoch):
    state = {'epoch': epoch+1,
             'state_dict': model.state_dict(),
             'optim_dict' : optimizer.state_dict()
            }
    torch.save(state,path_to_model)

def load_model(model,optimizer,path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    optimizer.load_state_dict(checkpoint['optim_dict'])
    epoch = checkpoint['epoch']
    return model,optimizer,epoch

def loss_fn_weightedloss(dic_instances):
    dic_instances=dic_instances.to(device)
    criterion = nn.CrossEntropyLoss(dic_instances)
    return criterion

def loss_fn():
    criterion = nn.CrossEntropyLoss()
    return criterion
    
def optimizer_fn():
    optimizer = optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=0.0001)
    return optimizer

def conf_mat_create(predicted,true,correct,total_30sec_epochs,conf_mat):
    total_30sec_epochs+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat=conf_mat+confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    return correct, total_30sec_epochs,conf_mat

def results_store_excel(correct_train,total_30sec_epochs_train,train_loss,correct_test,total_30sec_epochs_test,test_loss,epoch):
    avg_train_loss=train_loss/total_30sec_epochs_train
    avg_test_loss=test_loss/total_30sec_epochs_test
    accuracy_train=correct_train / total_30sec_epochs_train
    accuracy_test=correct_test / total_30sec_epochs_test
    lines=[epoch+1, avg_train_loss, accuracy_train, avg_test_loss, accuracy_test]
    out=open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    out.close()
    sheet1.append(lines)
    
def results_plot(df,df1):
    plt.plot(df[0],df[2],'-r',label='Train Accuracy CNN')
    plt.plot(df[0],df[4],'-b',label='Test Accuracy CNN')
    plt.plot(df1[0],df1[2],'-g',label='Train Accuracy CNNLSTM')
    plt.plot(df1[0],df1[4],'-y',label='Test Accuracy CNNLSTM')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,31))
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.show()

def train(model,data_iterator_train,data_iterator_test,batches_train,epochs):
    '''Training'''
    optimizer=optimizer_fn()
    early_stopping = EarlyStopping(path_to_model=path_to_model,patience=patience_epoch,verbose=True)
    if os.path.isfile(path_to_model):
        model,optimizer,start_epoch=load_model(model,optimizer,path_to_model)
        print(start_epoch)
    else:
        start_epoch=0
    if class_imbalance=='weightedcostfunc':
        lossfn=loss_fn_weightedloss(dic_train_instances)
    else:
        lossfn=loss_fn()
    for epoch in range(start_epoch,epochs):
        model.train()
        loss_train=0.0
        correct_train=0
        hidden_states_last=()
        conf_mat_train=np.zeros((5,5))
        total_30sec_epochs_train=0
        batch_no=0
        file_no=0
        print("checking")
        for train_idx, train_batch, train_labels in data_iterator_train:
            #print("batch:"+str(batch_no))
            train_batch, train_labels=train_batch.to(device), train_labels.to(device)
            if lstm_option==True:
                #print(train_idx)
                if train_idx[0]==f_file_length_dic_cumul[file_no] or train_idx[0]==0:
                    if train_idx[0]!=0:
                        file_no+=1
                    print("New Patient")
                    hidden_state=torch.zeros(2,1,20).to(device)
                    cell_state=torch.zeros(2,1,20).to(device)
                    initial_hidden_states=(hidden_state,cell_state)
                else:
                    initial_hidden_states=hidden_states_last
                if multiple_gpu==True:
                    train_batch_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(train_batch,seq_length)
                    train_labels_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(train_labels,seq_length)
                    train_batch_seq, train_labels_seq=train_batch_seq.to(device), train_labels_seq.to(device)
                    output_batch = model(train_batch_seq) # compute model output, loss and total train loss over one epoch
                else:
                    combined_input=[train_batch,initial_hidden_states]
                    output_batch, hidden_states_last = model(combined_input) # compute model output, loss and total train loss over one epoch
                
                pred = output_batch.argmax(dim=1, keepdim=True)
                train_labels_crop=train_labels.view(-1)
                if max(train_idx)!=train_idx[-1]:
                    idx_nump=train_idx.numpy()
                    max_idx=np.where(idx_nump==max(idx_nump))[0][0]
                    print(max_idx)
                    pred=pred[:max_idx+1]
                    train_labels_crop=train_labels_crop[:max_idx+1]
                    output_batch=output_batch[:max_idx+1]
            else:
                output_batch = model(train_batch) # compute model output, loss and total train loss over one epoch
                pred = output_batch.argmax(dim=1, keepdim=True)
                train_labels_crop=train_labels.view(-1)
            
            loss = lossfn(output_batch, train_labels_crop)
            loss_train+=(train_labels_crop.size()[0]*loss.item())
            optimizer.zero_grad()  # clear previous gradients, compute gradients of all variables wrt loss
            loss.backward()
            optimizer.step() # performs updates using calculated gradients
            
            #performance metrics of training dataset
            correct_train,total_30sec_epochs_train,conf_mat_train=conf_mat_create(pred, train_labels_crop, correct_train, total_30sec_epochs_train, conf_mat_train)
            print ('Epoch [{}/{}], Step [{}/{}], Loss: {:.4f}' .format(epoch+1, epochs, batch_no+1, batches_train, loss.item()))
            batch_no=batch_no+1
        
        correct_test,total_30sec_epochs_test,loss_test,conf_mat_test=validation(model, data_iterator_test, epoch)
        print("total 30sec epochs in the whole training data for one epoch of training and test:",total_30sec_epochs_train,total_30sec_epochs_test)
        results_store_excel(correct_train,total_30sec_epochs_train,loss_train,correct_test,total_30sec_epochs_test,loss_test,epoch)
        valid_loss=loss_test/total_30sec_epochs_test
        
        # early_stopping needs the validation loss to check if it has decresed, 
        # and if it has, it will make a checkpoint of the current model
        early_stopping(valid_loss,model,optimizer,epoch,conf_mat_train,conf_mat_test)
        if early_stopping.early_stop:
            print("Early stopping",epoch+1)
            break
    
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_train_best.tolist():
        sheet2.append(row)
    sheet2.append([0,1,2,3,4])
    for row in early_stopping.conf_mat_test_best.tolist():
        sheet2.append(row)
    print('Finished Training')
    #torch.load('checkpoint.pt'))
    #return model
    
def validation(model, data_iterator_val, epoch):
    """Testing"""
    model.eval()
    total_30sec_epochs=0
    val_loss = 0
    correct = 0
    s=0
    hidden_states_last=()
    conf_mat_test=np.zeros((5,5))
    file_no=0
    count_file_len=0
    if class_imbalance=='weightedcostfunc':
        lossfn1=loss_fn_weightedloss(dic_val_instances)
    else:
        lossfn1=loss_fn()
    for val_idx, val_data, val_labels in data_iterator_val:
        val_data, val_labels=val_data.to(device), val_labels.to(device)
        if lstm_option==True:
            if val_idx[0]==count_file_len:
                count_file_len+=file_length_dic_val[str(file_no)]
                file_no+=1
                print("New Patient")
                hidden_state=torch.zeros(2,1,20).to(device)
                cell_state=torch.zeros(2,1,20).to(device)
                initial_hidden_states=(hidden_state,cell_state)
            else:
                initial_hidden_states=hidden_states_last
            if multiple_gpu==True:
                val_data_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(val_data,seq_length)
                val_labels_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(val_labels,seq_length)
                val_data_seq, val_labels_seq=val_data_seq.to(device), val_labels_seq.to(device)
                output = model(val_data_seq)
            else:
                combined_input=[val_data,initial_hidden_states]
                output, hidden_states_last = model(combined_input)
            target_crop=val_labels.view(-1)
            val_pred = output.argmax(dim=1, keepdim=True) # get the index of the max log-probability
            if max(val_idx)!=val_idx[-1]:
                idx_nump=val_idx.numpy()
                max_idx=np.where(idx_nump==max(idx_nump))[0][0]
                print(max_idx)
                val_pred=val_pred[:max_idx+1]
                target_crop=target_crop[:max_idx+1]
                output=output[:max_idx+1]
        else:
            output = model(val_data)
            target_crop=val_labels.view(-1)
            val_pred = output.argmax(dim=1, keepdim=True) # get the index of the max log-probability
        
        s=s+val_data.shape[0]    
        #print("File name:",np.unique(fl_no_test.numpy()))
        print(val_data.shape[0])
        print(output.shape[0])
        loss1=lossfn1(output, target_crop).item()
        print("Test Loss:",loss1)
        val_loss += target_crop.size()[0]*loss1 # sum up batch loss
        correct,total_30sec_epochs,conf_mat_test=conf_mat_create(val_pred,target_crop,correct,total_30sec_epochs,conf_mat_test)
    
    print("conf_mat_test:",conf_mat_test)
    print("total_30sec_epochs:",total_30sec_epochs)
    print("s:",s)
    print('\nTest set: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%), Epoch:{}\n'.format(
        val_loss/total_30sec_epochs, correct, total_30sec_epochs,
        100. * correct / total_30sec_epochs,epoch+1))
    return correct,total_30sec_epochs,val_loss,conf_mat_test

if __name__ == '__main__': 
    # Initialization
    random.seed(30)
    classes=[0,1,2,3,4]
    args=args_parameters()
    batch_size=args.batch_size
    n_workers=args.n_workers
    time_steps=args.time_steps
    max_epochs = args.max_epochs
    n_channels= args.n_channels
    seq_length= args.seq_length
    lstm_option=args.lstm_option #mention this parameter as true if you want to add lstm layer to the model
    rc_option= args.rc_option
    class_imbalance= args.class_imbalance
    lr=args.learning_rate
    modality_pipelines=args.modality_pipelines
    patience_epoch=args.patience_epoch
    
    if torch.cuda.device_count() > 1:
        multiple_gpu=True
    else:
        multiple_gpu=False
    
    # CUDA for PyTorch
    use_cuda = torch.cuda.is_available()
    device = torch.device("cuda:0" if use_cuda else "cpu")
    
    
    ####################  train  ################
    path_to_hdf5_file_train='D:/DEEPSLEEP/test_for_github/datasets/shhs/train/hdf5_file_train_all_chunking_shhs1.hdf5'
    model_name=input('Enter the name of the model (without the extension .tar):')
    path_to_model="D:/DEEPSLEEP/test_for_github/saved_models/shhs/"+model_name+".tar"
    
    path_to_file_length_train='D:/DEEPSLEEP/test_for_github/datasets/shhs/train/trainFilesNum30secEpochs_all_shhs1.pkl'
    f_file_length_train=open(path_to_file_length_train,'rb')
    file_length_dic_train=pickle.load(f_file_length_train)
    batches_train=np.sum(np.ceil(np.array(list(file_length_dic_train.values()))/batch_size),dtype='int32')
    f_file_length_train.close()
    
    path_to_file_length_cumul='D:/DEEPSLEEP/test_for_github/datasets/shhs/train/trainFilesNum30secEpochsCumulative_all_shhs1.pkl'
    f_file_length_cumul=open(path_to_file_length_cumul,'rb')
    f_file_length_dic_cumul=pickle.load(f_file_length_cumul)
    ##################  train end  #################
    
    
    #################   val       ################
    path_to_hdf5_file_val ='D:/DEEPSLEEP/test_for_github/datasets/shhs/val/hdf5_file_val_all_chunking_shhs1.hdf5'
    result_filename=input('Enter the name of the result file (without the extension):')
    path_to_results="D:/DEEPSLEEP/test_for_github/results/shhs/"+result_filename+".xlsx"
    path_to_results_text="D:/DEEPSLEEP/test_for_github/results/shhs/"+result_filename+".txt"
    
    path_to_file_length_val='D:/DEEPSLEEP/test_for_github/datasets/shhs/val/valFilesNum30secEpochs_all_shhs1.pkl'
    f_file_length_val=open(path_to_file_length_val,'rb')
    file_length_dic_val=pickle.load(f_file_length_val)
    batches_val=np.sum(np.ceil(np.array(list(file_length_dic_val.values()))/batch_size),dtype='int32')
    f_file_length_val.close()
    ###############    val end   ################# 
    
    #Weight of class instances for weighted cost function
    if class_imbalance=='weightedcostfunc1':
        dic_train_instances=utils.class_distribution_weightedloss1(path_to_hdf5_file_train)
        dic_val_instances=utils.class_distribution_weightedloss1(path_to_hdf5_file_val)
    elif class_imbalance=='weightedcostfunc2':
        dic_train_instances=utils.class_distribution_weightedloss2(path_to_hdf5_file_train)
        dic_val_instances=utils.class_distribution_weightedloss2(path_to_hdf5_file_val)
    
    # set file path
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        #sheet1 = wb.get_sheet_by_name('epoch accuracy-loss train_val')
        sheet2 = wb.get_sheet_by_name('confusion matrix train_val')
        sheet3 = wb.get_sheet_by_name('confusion matrix test')
    else:
        wb=Workbook()
        sheet1=wb.active
        header=['Epoch','Avg Loss Train','Accuracy Train','Avg Loss Test','Accuracy Test']
        sheet1.append(header)
        sheet2=wb.create_sheet('confusion matrix train_val')
        sheet3 = wb.create_sheet('confusion matrix test')
    
    #model setup
    model=models.DeepSleepSpatialTemporalNet(modality_pipelines, lstm_option, rc_option, multiple_gpu)
    for name, param in model.named_parameters():
        if param.requires_grad:
            print(name)
    if torch.cuda.device_count() > 1:
        print("Let's use", torch.cuda.device_count(), "GPUs!")
        model = nn.DataParallel(model)
    if lstm_option==True:
        #training preparation
        pretrained_model=input('Enter the name of the pretrained ST (CNN) model (without the extension):')
        path_to_pretrained_model="D:/DEEPSLEEP/test_for_github/saved_models/shhs/"+pretrained_model+".tar"
        model=load_pretrained_model_for_LSTM(model,path_to_pretrained_model)
        #if you want to freeze some layers of cnn and train the other layers, then use this function 
        #model=utils.freeze_layers(model)
        
        
        print("start generator train")
        data_gen_train=utils.my_generator1(path_to_hdf5_file_train)  
        print("start dataloader train")
        #this is for sequence learning using LSTM for training
        sampler_train=SequentialSampler(data_gen_train)
        data_iterator_Train=DataLoader(data_gen_train,batch_size=1,num_workers=n_workers,batch_sampler=utils.CustomSequentialLSTMBatchSampler_ReturnAllChunks(sampler_train,batch_size,file_length_dic_train,seq_length))
        
        
        #testing preparation
        print("start generator test")
        data_gen_val=utils.my_generator1(path_to_hdf5_file_val)
        print("start dataloader test")
        #this is for sequence learning using LSTM for test
        sampler_val=SequentialSampler(data_gen_val)
        data_iterator_val=DataLoader(data_gen_val,batch_size=1,num_workers=n_workers,batch_sampler=utils.CustomSequentialLSTMBatchSampler_ReturnAllChunks(sampler_val,batch_size,file_length_dic_val,seq_length))
    
    else:
        #training preparation
        print("start generator train")
        data_gen_train=utils.my_generator1(path_to_hdf5_file_train)  
        print("start dataloader train")
        sampler=utils.CustomRandomSamplerSlicedShuffled(path_to_hdf5_file_train,f_file_length_dic_cumul)
        if class_imbalance=='oversampling':
            batch_sampler_oversampling=utils.CustomWeightedRandomBatchSamplerSlicedShuffled(sampler,batch_size,f_file_length_dic_cumul)
            data_iterator_Train=DataLoader(data_gen_train, batch_size=1, num_workers=n_workers, batch_sampler=batch_sampler_oversampling)
        else:
            batch_sampler_random_shuffling=utils.CustomRandomBatchSamplerSlicedShuffled(sampler,batch_size,file_length_dic_train)
            data_iterator_Train=DataLoader(data_gen_train, batch_size=1, num_workers=n_workers, batch_sampler=batch_sampler_random_shuffling)
        
        #test preparation
        print("start generator test")
        data_gen_val=utils.my_generator1(path_to_hdf5_file_val)
        print("start dataloader test")
        data_iterator_val=DataLoader(data_gen_val,batch_size=batch_size,num_workers=n_workers)
    model.to(device)
    for name, param in model.named_parameters():
        if param.requires_grad:
            print(name)
    #training
    train(model,data_iterator_Train,data_iterator_val,batches_train,max_epochs)
    
    #testing
    #path_to_trained_model="C:/Users/PathakS/OneDrive - Universiteit Twente/deepsleep/final_codes/SHHS/saved_models/cnnlstm/model2d_shhs1_4692_5213_allCh_weightedloss1_cnnlstm_ReturnAllChunk.tar"
    #model=load_trained_model_for_testing(model,path_to_trained_model)
    #test(model, data_iterator_Test)
    
    #save file
    wb.save(path_to_results)