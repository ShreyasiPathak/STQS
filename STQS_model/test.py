# -*- coding: utf-8 -*-
"""
Created on Tue Jun 18 18:28:49 2019

@author: PathakS
"""

import os
import h5py
import torch
import pickle
import random
import argparse
import numpy as np
import torch.nn as nn
import openpyxl as op
from random import shuffle
import torch.optim as optim
from openpyxl import Workbook
import matplotlib.pyplot as plt
from collections import OrderedDict
from pytorchtools import EarlyStopping
from torch.utils.data import DataLoader, SequentialSampler
from sklearn.metrics import confusion_matrix

import utils
import models

def args_parameters():
    parser = argparse.ArgumentParser()
    parser.add_argument('--batch_size', type=int, default=192)
    parser.add_argument('--n_workers', type=int, default=3)
    parser.add_argument('--learning_rate', type=float, default=0.0001)
    parser.add_argument('--time_steps', type=int, default=3750)
    parser.add_argument('--n_channels', type=int, default=5)
    parser.add_argument('--modality_pipelines', type=int, default=3)
    parser.add_argument('--seq_length', type=int, default=8)
    parser.add_argument('--lstm_option', type=bool, default=False)
    parser.add_argument('--rc_option', type=bool, default=False)
    args = parser.parse_args()
    return args

def load_trained_model_for_testing(model, path):
    model_parameters=torch.load(path)
    model_weights=model_parameters['state_dict']
    state_dict_remove_module = OrderedDict()
    for k, v in model_weights.items():
        if k[:7]=='module.':
            print(k, v.shape) #k coming from pretrained model
            name = k[7:] # remove `module.`
            state_dict_remove_module[name] = v
        else:
            state_dict_remove_module[k] = v
    model.load_state_dict(state_dict_remove_module)
    return model.to(device)

def load_pretrained_model_for_LSTM(model,path):
    state_dict_new_model=model.state_dict()
    checkpoint = torch.load(path)
    state_dict_pretrained=checkpoint['state_dict']
    state_dict_remove_module = OrderedDict()
    for k, v in state_dict_pretrained.items():
        if k[:7]=='module.':
            if k!='module.linear.weight' and k!='module.linear.bias':
                name = k[7:] # remove `module.`
                state_dict_remove_module[k] = v
        else:
            if k!='linear.weight' and k!='linear.bias':
                name = 'module.' + k # remove `module.`
                state_dict_remove_module[name] = v
    state_dict_new_model.update(state_dict_remove_module)
    model.load_state_dict(state_dict_new_model)
    return model

def load_model(model,optimizer,path):
    checkpoint = torch.load(path)
    model.load_state_dict(checkpoint['state_dict'])
    optimizer.load_state_dict(checkpoint['optim_dict'])
    epoch = checkpoint['epoch']
    return model,optimizer,epoch

def loss_fn_weightedloss(dic_instances):
    dic_instances=dic_instances.to(device)
    criterion = nn.CrossEntropyLoss(dic_instances)
    return criterion

def loss_fn():
    criterion = nn.CrossEntropyLoss()
    return criterion
    
def optimizer_fn():
    optimizer = optim.Adam(filter(lambda p: p.requires_grad, model.parameters()), lr=0.0001)
    return optimizer

def conf_mat_create(predicted,true,correct,total_30sec_epochs,conf_mat):
    total_30sec_epochs+=true.size()[0]
    correct += predicted.eq(true.view_as(predicted)).sum().item()
    conf_mat=conf_mat+confusion_matrix(true.cpu().numpy(),predicted.cpu().numpy(),labels=classes)
    return correct, total_30sec_epochs,conf_mat

def results_store_excel(correct_train,total_30sec_epochs_train,train_loss,correct_test,total_30sec_epochs_test,test_loss,epoch):
    avg_train_loss=train_loss/total_30sec_epochs_train
    avg_test_loss=test_loss/total_30sec_epochs_test
    accuracy_train=correct_train / total_30sec_epochs_train
    accuracy_test=correct_test / total_30sec_epochs_test
    lines=[epoch+1, avg_train_loss, accuracy_train, avg_test_loss, accuracy_test]
    out=open(path_to_results_text,'a')
    out.write(str(lines)+'\n')
    out.close()
    sheet1.append(lines)
    
def results_plot(df,df1):
    plt.plot(df[0],df[2],'-r',label='Train Accuracy CNN')
    plt.plot(df[0],df[4],'-b',label='Test Accuracy CNN')
    plt.plot(df1[0],df1[2],'-g',label='Train Accuracy CNNLSTM')
    plt.plot(df1[0],df1[4],'-y',label='Test Accuracy CNNLSTM')
    plt.legend(loc='upper left')
    plt.xticks(np.arange(1,31))
    plt.xlabel('Epochs')
    plt.ylabel('Accuracy')
    plt.show()

def test(model, data_iterator_test):
    """Testing"""
    model.eval()
    total_30sec_epochs=0
    test_loss = 0
    correct = 0
    s=0
    hidden_states_last=()
    batch_no=0
    conf_mat_test=np.zeros((5,5))
    file_no=0
    count_file_len=0
    #lossfn1=loss_fn_weightedloss(dic_test_instances)
    lossfn1=loss_fn()
    for test_idx, test_data, test_labels in data_iterator_test:
        test_data, test_labels=test_data.to(device), test_labels.to(device)
        if lstm_option==True:
            if test_idx[0]==count_file_len:
                count_file_len+=file_length_dic_test[str(file_no)]
                file_no+=1
                print("New Patient")
                hidden_state=torch.zeros(2,1,20).to(device)
                cell_state=torch.zeros(2,1,20).to(device)
                initial_hidden_states=(hidden_state,cell_state)
            else:
                initial_hidden_states=hidden_states_last
            if multiple_gpu==True:
                test_data_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(test_data,seq_length)
                test_labels_seq=utils.to_sequence_for_gpu_LSTM_ReturnAllChunk(test_labels,seq_length)
                test_data_seq, test_labels_seq=test_data_seq.to(device), test_labels_seq.to(device)
                output = model(test_data_seq)
            else:
                combined_input=[test_data,initial_hidden_states]
                output, hidden_states_last = model(combined_input)
            target_crop=test_labels.view(-1)
            test_pred = output.argmax(dim=1, keepdim=True) # get the index of the max log-probability
            if max(test_idx)!=test_idx[-1]:
                idx_nump=test_idx.numpy()
                max_idx=np.where(idx_nump==max(idx_nump))[0][0]
                print(max_idx)
                test_pred=test_pred[:max_idx+1]
                target_crop=target_crop[:max_idx+1]
                output=output[:max_idx+1]
        else:
            output = model(test_data)
            target_crop=test_labels.view(-1)
            test_pred = output.argmax(dim=1, keepdim=True) # get the index of the max log-probability
        
        s=s+test_data.shape[0]    
        batch_no=batch_no+1
        print(test_data.shape[0])
        print(output.shape[0])
        loss1=lossfn1(output, target_crop).item()
        print("\nTest Loss for {}/{} is {:.4f}:\n".format(batch_no,batches_test,loss1))
        test_loss += target_crop.size()[0]*loss1 # sum up batch loss
        correct,total_30sec_epochs,conf_mat_test=conf_mat_create(test_pred,target_crop,correct,total_30sec_epochs,conf_mat_test)
    
    print("conf_mat_test:",conf_mat_test)
    print("total_30sec_epochs:",total_30sec_epochs)
    print("s:",s)
    print('\nTest set: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
        test_loss/total_30sec_epochs, correct, total_30sec_epochs,
        100. * correct / total_30sec_epochs))
    sheet3.append([0,1,2,3,4])
    for row in conf_mat_test.tolist():
        sheet3.append(row)
    metrics_class,metrics_model=utils.metrics_confusion_matrix_per_class(conf_mat_test)
    print(metrics_class)
    print(metrics_model)
    sheet3.append(['Class','Prec','Rec','F1'])
    for i,res in enumerate(metrics_class):
        sheet3.append([classes_name[i]]+res)
    sheet3.append(['Acc','Bal_Acc','Macro_F1','Cohens Kappa','Wt_Macro_F1'])
    sheet3.append(metrics_model)

if __name__ == '__main__': 
    # Initialization
    random.seed(30)
    classes=[0,1,2,3,4]
    classes_name=['W','N1','N2','N3','REM']
    args=args_parameters()
    batch_size=args.batch_size
    n_workers=args.n_workers
    time_steps=args.time_steps
    n_channels= args.n_channels
    seq_length= args.seq_length
    lstm_option=args.lstm_option #mention this parameter as true if you want to add lstm layer to the model
    rc_option= args.rc_option
    lr=args.learning_rate
    modality_pipelines=args.modality_pipelines
    if torch.cuda.device_count() > 1:
        multiple_gpu=True
    else:
        multiple_gpu=False
    
    # CUDA for PyTorch
    use_cuda = torch.cuda.is_available()
    device = torch.device("cuda:0" if use_cuda else "cpu")
    
    #################   test       ################
    path_to_hdf5_file_test='D:/DEEPSLEEP/test_for_github/datasets/shhs/test/hdf5_file_test_all_chunking_shhs1.hdf5'
    result_filename=input('Enter the name of the result file created during the training (without the extension):')
    path_to_results="D:/DEEPSLEEP/test_for_github/results/shhs/"+result_filename+".xlsx"
    path_to_results_text="D:/DEEPSLEEP/test_for_github/results/shhs/"+result_filename+".txt"
    
    path_to_file_length_test='D:/DEEPSLEEP/test_for_github/datasets/shhs/test/testFilesNum30secEpochs_all_shhs1.pkl'
    f_file_length_test=open(path_to_file_length_test,'rb')
    file_length_dic_test=pickle.load(f_file_length_test)
    batches_test=np.sum(np.ceil(np.array(list(file_length_dic_test.values()))/batch_size),dtype='int32')
    f_file_length_test.close()
    ###############    test end   ################# 
    
    #Weight of class instances for weighted cost function
    #dic_train_instances=utils.class_distribution(path_to_hdf5_file_train)
    #dic_test_instances=utils.class_distribution(path_to_hdf5_file_test)
    
    # set file path
    if os.path.isfile(path_to_results):
        wb = op.load_workbook(path_to_results)
        sheet3 = wb.get_sheet_by_name('confusion matrix test')
    else:
        wb=Workbook()
        sheet1=wb.active
        header=['Epoch','Avg Loss Train','Accuracy Train','Avg Loss Test','Accuracy Test']
        sheet1.append(header)
        sheet2=wb.create_sheet('confusion matrix train_val')
        sheet3 = wb.create_sheet('confusion matrix test')
    
    #model setup
    model=models.DeepSleepSpatialTemporalNet(modality_pipelines, lstm_option, rc_option, multiple_gpu)
    if torch.cuda.device_count() > 1:
        print("Let's use", torch.cuda.device_count(), "GPUs!")
        model = nn.DataParallel(model)
    
    if lstm_option==True:
        #testing preparation
        print("start generator test")
        data_gen_test=utils.my_generator1(path_to_hdf5_file_test)
        print("start dataloader test")
        #this is for sequence learning using LSTM for test
        sampler_test=SequentialSampler(data_gen_test)
        data_iterator_Test=DataLoader(data_gen_test,batch_size=1,num_workers=n_workers,batch_sampler=utils.CustomSequentialLSTMBatchSampler_ReturnAllChunks(sampler_test,batch_size,file_length_dic_test,seq_length))
    
    else:
        #test preparation
        print("start generator test")
        data_gen_test=utils.my_generator1(path_to_hdf5_file_test)
        print("start dataloader test")
        data_iterator_Test=DataLoader(data_gen_test,batch_size=batch_size,num_workers=n_workers)
    
    model.to(device)
    
    #testing
    model_name=input('Enter the name of the model (without the extension .tar):')
    path_to_trained_model="D:/DEEPSLEEP/test_for_github/saved_models/shhs/"+model_name+".tar"
    model=load_trained_model_for_testing(model,path_to_trained_model)
    test(model, data_iterator_Test)
    
    #save file
    wb.save(path_to_results)